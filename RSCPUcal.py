def RSCPU_count(list):
    import openpyxl
    f=5
    g=0
    wb = openpyxl.load_workbook('RSCPU_template.xlsx')
    ws = wb.active
    for e in list:

        wb_number = openpyxl.load_workbook(e + '.xlsx')
        ws_numner = wb_number.active

        ###  prepare all AApair
        AA_pair = []
        AA = ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'V', 'W', 'Y']
        # print(len(AA))
        for a in range(0, 20):
            for b in range(0, 20):
                AA_pair.append(AA[a] + AA[b])
        # print(AA_pair)
        # print(len(AA_pair))

        ### Collect values from statistics
        for c in range(2, 3723):
            for d in range(2, 3723):
                s_codon = ws.cell(row=d, column=4).value
                f_codon = ws_numner.cell(row=c, column=1).value
                f_number = ws_numner.cell(row=c, column=2).value
                if f_codon == s_codon:
                    ws_numner.cell(row=d, column=3, value=f_number)

            ### Collect table information
        AA_pair_all = []
        code_pos1 = []
        code_pos2 = []
        number = []

        for i in range(2, 3723):
            AA_pair_e = ws.cell(row=i, column=3).value
            AA_pair_all.append(AA_pair_e)

            code_pos1_e = ws.cell(row=i, column=1).value
            code_pos1.append(code_pos1_e)
            code_pos2_e = ws.cell(row=i, column=2).value
            code_pos2.append(code_pos2_e)

            number_e = ws_numner.cell(row=i, column=3).value
            number.append(number_e)

        # print(AA_pair_all)
        # print(len(AA_pair_all))
        # print(code_pos2)
        # print(code_pos1)
        # print(len(number))

        ### count n(i)
        RSCPU_all = ['0', '0']
        ### count X(i)
        for j in range(0, 3721):
            X = int(number[j])
            N = 0
            n = 0
            AA_pair_X = AA_pair_all[j]
            for z in range(0, 3721):
                if str(AA_pair_all[z]) == AA_pair_X:
                    N = N + int(number[z])
                    n = n + 1
            RSCPU = X * n / N
            RSCPU_all.append(RSCPU)


        for t in range(2, 3723):
            ws.cell(row=t, column=f, value=RSCPU_all[t])
        ws.cell(row=1, column=f, value=e)
        f=f+1
        g=g+1

        print(g)

    wb.save('new_RSCPU.xlsx')









