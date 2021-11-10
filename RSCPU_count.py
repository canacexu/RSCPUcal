### Count the frequency of codon occurrence by reading the CDS sequence.
def RSCPU_s(name):
    for t in name:
        import openpyxl

        wb_n = openpyxl.load_workbook('zero.xlsx')
        ws_n = wb_n.active

        # Extract gene sequence
        GeneDetail = ''
        AllGeneDetails = []
        with open(t + '.txt', 'r+') as fileStream:
            while True:
                line = fileStream.readline()
                # Read a line at once

                if not line:
                    AllGeneDetails.append(GeneDetail)
                    GeneDetail = ''
                    break
                if line[0] == '\n':
                    continue
                # check if a line is the last line

                if line[0] in ('A', 'T', 'G', 'C'):
                    if line[-1] == '\n':
                        GeneDetail = GeneDetail + line[:-1]
                    #                GeneDetail= GeneDetail[0:-3]
                    else:
                        print('fix: last line')
                        GeneDetail = GeneDetail + line
                #        GeneDetail = GeneDetail[0:-3]
                if line[0] == '>':
                    AllGeneDetails.append(GeneDetail)

                    GeneDetail = ''
                    # Clear the buffer for the NEXT GeneDetail



        # Get codon

        def DNA_complement(sequence):
            # sequence = sequence.upper()
            sequence = sequence.replace('A', 'u')
            sequence = sequence.replace('T', 'a')
            sequence = sequence.replace('C', 'g')
            sequence = sequence.replace('G', 'c')

            return sequence.upper()

        for Index in range(len(AllGeneDetails)):
            AllGeneDetails[Index] = DNA_complement(AllGeneDetails[Index])

        # form codon-pair and count
        codon_pair = []
        for k in range(len(AllGeneDetails)):
            for j in range(2, len(AllGeneDetails[k]), 3):
                codon_pair.append(AllGeneDetails[k][j - 2:j + 1] + '-' + AllGeneDetails[k][j + 1:j + 4])
            codon_pair = codon_pair[0:-1]
        d = {}
        for pair in codon_pair:
            d[pair] = d.get(pair, 0) + 1
        ls = list(d.items())
        # print(len(ls))
        # Fill the statistics in the table
        ws_n.cell(row=1, column=1, value="codon-pair")
        for i in range(2, len(ls) + 2):
            for j in range(2,3723):
                c=ws_n.cell(row=j, column=1).value
                # ws_n.cell(row=j,column=2,value=0)
                d=ls[i - 2][0]
                if d==c:
                    ws_n.cell(row=j, column=1, value=ls[i - 2][0])
                    ws_n.cell(row=j, column=2, value=ls[i - 2][1])


        print('FINISH_'+t)

        wb_n.save(t + '.xlsx')


